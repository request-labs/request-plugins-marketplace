#!/usr/bin/env python3
"""Generate Excel invoice report from a JSON file.

Usage:
    python3 gen_excel.py input.json output.xlsx

The agent should:
1. Write the JSON data to a temp file using the Write tool (no permission prompt)
2. Run: python3 gen_excel.py /tmp/invoices.json /path/to/output.xlsx
3. The script auto-deletes the input JSON after reading it.

Input JSON schema:
{
  "faturas": [
    {"ficheiro": "x.pdf", "fornecedor": "X", "nif_fornecedor": "123",
     "numero": "FT1", "data_emissao": "01-01-2026", "total": 100.0,
     "moeda": "EUR", "confidence": 1.0}
  ],
  "resumo": {
    "total": 10, "sucesso": 7, "baixa_confianca": 1,
    "sem_parser": 1, "erros": 1
  }
}
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")

COLUMNS = [
    ("Ficheiro", 42),
    ("Fornecedor", 30),
    ("NIF", 15),
    ("Nº Fatura", 30),
    ("Data", 12),
    ("Total", 12),
    ("Moeda", 8),
    ("Confidence", 12),
]

FIELDS = ["ficheiro", "fornecedor", "nif_fornecedor", "numero", "data_emissao", "total", "moeda", "confidence"]


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def build_faturas_sheet(wb, faturas):
    ws = wb.active
    ws.title = "Faturas"

    # Header
    for i, (name, width) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=i, value=name)
        ws.column_dimensions[get_column_letter(i)].width = width

    style_header(ws, len(COLUMNS))

    # Data rows
    for row_idx, inv in enumerate(faturas, 2):
        for col_idx, field in enumerate(FIELDS, 1):
            ws.cell(row=row_idx, column=col_idx, value=inv.get(field))

    # Total column number format
    total_col = FIELDS.index("total") + 1
    for row_idx in range(2, len(faturas) + 2):
        ws.cell(row=row_idx, column=total_col).number_format = "#,##0.00"

    # SUM row
    sum_row = len(faturas) + 2
    ws.cell(row=sum_row, column=total_col - 1, value="TOTAL").font = Font(bold=True)
    total_letter = get_column_letter(total_col)
    ws.cell(
        row=sum_row,
        column=total_col,
        value=f"=SUM({total_letter}2:{total_letter}{sum_row - 1})",
    ).number_format = "#,##0.00"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(faturas) + 1}"

    return ws


def build_resumo_sheet(wb, resumo):
    ws = wb.create_sheet("Resumo")

    ws.cell(row=1, column=1, value="Métrica")
    ws.cell(row=1, column=2, value="Valor")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    style_header(ws, 2)

    rows = [
        ("Total ficheiros", resumo.get("total", 0)),
        ("Sucesso", resumo.get("sucesso", 0)),
        ("Baixa confiança", resumo.get("baixa_confianca", 0)),
        ("Sem parser", resumo.get("sem_parser", 0)),
        ("Erros", resumo.get("erros", 0)),
    ]

    for i, (label, value) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    return ws


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 gen_excel.py input.json output.xlsx", file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path) as f:
        data = json.load(f)

    # Auto-cleanup input JSON
    os.remove(input_path)

    wb = Workbook()
    build_faturas_sheet(wb, data.get("faturas", []))
    build_resumo_sheet(wb, data.get("resumo", {}))

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
